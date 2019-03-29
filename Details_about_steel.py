import openpyxl
import sys
from PyQt5 import QtCore
from PyQt5.QtWidgets import *
from PyQt5.QtGui import *
from PyQt5.QtCore import *
import sqlite3
from sqlite3 import Error

class Item_Detail(QWidget):

    def __init__(self,row):
        super().__init__()
        self.title = "Details"
        self.row = row
        self.initUI()

    @pyqtSlot()
    def select_all_tasks(self,conn):
        """
        Query all rows in the tasks table
        :param conn: the Connection object
        :return:
        """
        cur = conn.cursor()
        cur.execute("SELECT * FROM Angles WHERE Id="+str(self.row)+";")

        rows = cur.fetchall()
        rows = list(rows[0])

        self.specified_row = rows
        print(self.specified_row)
        print(len(self.specified_row))

    @pyqtSlot()
    def create_connection(self,db_file):
        """ create a database connection to the SQLite database
            specified by the db_file
        :param db_file: database file
        :return: Connection object or None
        """
        try:
            conn = sqlite3.connect(db_file)
            return conn
        except Error as e:
            print(e)

        return None

    def db_for_detail(self):
        # create a database connection
        conn = self.create_connection("steel_sections.sqlite")
        with conn:
            self.select_all_tasks(conn)


    def initUI(self):
        self.setWindowTitle(self.title)
        self.showFullScreen()
        # self.Fetch_details()

        # Set window background color
        self.setAutoFillBackground(True)
        p = self.palette()
        p.setColor(self.backgroundRole(), Qt.white)
        self.setPalette(p)

        self.db_for_detail()


        Vertical_layout = QVBoxLayout()

        Vertical_layout_1 = QVBoxLayout()
        Vertical_layout_1.setSpacing(20)

        Vertical_layout_2 = QVBoxLayout()
        Vertical_layout_2.setSpacing(20)

        Vertical_layout_1.setContentsMargins(150, 80, 150, 200)
        Vertical_layout_2.setContentsMargins(150, 80, 150, 200)

        Horizontal_layout = QHBoxLayout()
        Horizontal_layout_Main_heading = QHBoxLayout()

        Window_heading = QLabel("Details of : "+self.specified_row[1],self)
        Horizontal_layout_Main_heading.addWidget(Window_heading)

        Horizontal_layout_Main_heading.setAlignment(Qt.AlignCenter)
        Horizontal_layout_Main_heading.setContentsMargins(0,0,0,0)

        # Vertical_layout_1.addWidget(Window_heading)
        self.details_headline=["Id","Designation","Mass","Area","AXB","t","FlangeSlope","R1","R2","Cz","Cy","Tan","Iz","Iy","Iu(max)","Iv(min)","rz","ry","ru(max)","rv(min)","Zz","Zy","Zpz","Zpy","Source"]
        print(len(self.details_headline))
        Horizontal_layout_End_Button = QHBoxLayout()

        Go_to_home_Button = QPushButton('Go to Home', self)
        Horizontal_layout_End_Button.addWidget(Go_to_home_Button)

        Horizontal_layout_End_Button.setAlignment(Qt.AlignCenter)
        Horizontal_layout_End_Button.setContentsMargins(0,0,0,200)
        for i in range(int(len(self.specified_row)/2)):

            Horizontal_layout_1 = QHBoxLayout()
            Horizontal_layout_1.setSpacing(5)
            # Horizontal_layout_1.setContentsMargins(100, 100, 100, 100)

            heading = QLabel(str(self.details_headline[i])+" = ",self)
            value = QLabel(str(self.specified_row[i]),self)
            # print(str(self.details_headline[i]),str(self.specified_row[i]))

            Horizontal_layout_1.addWidget(heading)
            Horizontal_layout_1.addWidget(value)

            Horizontal_layout_2 = QHBoxLayout()
            Horizontal_layout_2.setSpacing(5)


            heading = QLabel(str(self.details_headline[i+int(len(self.specified_row)/2)])+" = ",self)
            value = QLabel(str(self.specified_row[i+int(len(self.specified_row)/2)]),self)
            # print(i+int(len(self.specified_row)/2))
            #
            Horizontal_layout_2.addWidget(heading)
            Horizontal_layout_2.addWidget(value)

            Vertical_layout_1.addLayout(Horizontal_layout_1)
            Vertical_layout_2.addLayout(Horizontal_layout_2)

        Horizontal_layout.addLayout(Vertical_layout_1)
        Horizontal_layout.addLayout(Vertical_layout_2)
        Vertical_layout.addLayout(Horizontal_layout_Main_heading)
        Vertical_layout.addLayout(Horizontal_layout)
        Vertical_layout.addLayout(Horizontal_layout_End_Button)


        self.setLayout(Vertical_layout)

        # print(self.details_headline,self.steel_details)
        self.show()

    # def Fetch_details(self):
    #
    #     self.path = "new_sections.xlsx"
    #     wb_object = openpyxl.load_workbook(self.path)
    #
    #     sheet_obj = wb_object.active
    #     m_col = sheet_obj.max_column
    #
    #     self.details_headline=[]
    #
    #     for i in range(1,m_col+1):
    #         cell_obj2 = sheet_obj.cell(row =1,column=i)
    #
    #         details_headline_value = cell_obj2.value
    #         self.details_headline.append(details_headline_value)
